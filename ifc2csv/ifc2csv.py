# (CC0) balaji.work
# Updated to have any paramters exported from IFC Models
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import ifcopenshell
import csv
import os
import sys
import pandas as pd
import webbrowser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime  # For timestamp in Excel filename

# Log errors to a file when running as an exe
def log_errors_to_file(log_file="error_log.txt"):
    sys.stdout = open(log_file, "w")
    sys.stderr = sys.stdout

if getattr(sys, 'frozen', False):
    log_errors_to_file()

# Global variables
selected_properties = set()
available_properties = []
parameter_checkboxes = []  # Initialize the list of checkboxes

# Load user-defined properties from txt file
def load_selected_properties(file_path):
    global selected_properties
    try:
        with open(file_path, "r") as f:
            selected_properties = {line.strip() for line in f if line.strip()}
        print(f"Loaded {len(selected_properties)} properties from {file_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Error reading properties file: {e}")


# Extract properties from IFC file
def extract_ifc_properties(ifc_file_path, output_csv_path):
    global available_properties
    print(f"Processing: {ifc_file_path}")
    ifc_file = ifcopenshell.open(ifc_file_path)
    elements = ifc_file.by_type("IfcElement")

    element_data = []
    all_columns = set()

    for element in elements:
        element_id = element.GlobalId
        element_name = element.Name if element.Name else "Unknown"
        element_type = element.is_a()
        properties = {"GlobalId": element_id, "Name": element_name, "Type": element_type}

        if hasattr(element, "IsDefinedBy"):
            for rel in element.IsDefinedBy:
                if rel.is_a("IfcRelDefinesByProperties"):
                    prop_set = rel.RelatingPropertyDefinition
                    if hasattr(prop_set, "HasProperties"):
                        for prop in prop_set.HasProperties:
                            if hasattr(prop, "Name") and hasattr(prop, "NominalValue"):
                                prop_name = prop.Name
                                prop_value = prop.NominalValue.wrappedValue
                                properties[prop_name] = prop_value
                                all_columns.add(prop_name)

        element_data.append(properties)

    # Store available properties for checkbox selection
    available_properties = sorted(all_columns)

    # Prepare final column order with dynamically added columns
    final_columns = ["GlobalId", "Name", "Type"] + sorted(all_columns)

    # Write to CSV
    with open(output_csv_path, mode="w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=final_columns)
        writer.writeheader()
        for data in element_data:
            writer.writerow(data)

    print(f"Saved CSV: {output_csv_path}")


# Process all IFC files in a directory
def process_ifc_directory(input_dir, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    files_to_process = [f for root, _, files in os.walk(input_dir) for f in files if f.lower().endswith(".ifc")]
    total_files = len(files_to_process)

    if total_files == 0:
        print("No IFC files found.")
        return

    file_count = 0
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.lower().endswith(".ifc"):
                file_count += 1
                input_file_path = os.path.join(root, file_name)
                relative_path = os.path.relpath(root, input_dir)
                output_subdir = os.path.join(output_dir, relative_path)
                os.makedirs(output_subdir, exist_ok=True)
                output_file_name = os.path.splitext(file_name)[0] + ".csv"
                output_file_path = os.path.join(output_subdir, output_file_name)

                # Update progress bar and terminal output
                progress_var.set((file_count / total_files) * 100)
                progress_label.config(text=f"Processing file {file_count} of {total_files}: {file_name}")
                app.update_idletasks()

                print(f"Processing file {file_count} of {total_files}: {file_name}")
                try:
                    extract_ifc_properties(input_file_path, output_file_path)
                except Exception as e:
                    print(f"Error processing {input_file_path}: {str(e)}")

    # After processing all IFC files, load the checkboxes for parameters
    load_checkboxes()


# Combine all CSVs and create a validation Excel file
def create_combined_excel(output_dir):
    global selected_properties
    combined_data = []
    all_columns = set()

    # Loop through all CSV files and load them into a dataframe
    for root, _, files in os.walk(output_dir):
        for file_name in files:
            if file_name.lower().endswith(".csv"):
                csv_file_path = os.path.join(root, file_name)
                df = pd.read_csv(csv_file_path)

                # If selected_properties is not empty, filter the dataframe based on selected columns
                if selected_properties:
                    # Include only the columns that are selected by the user
                    df = df[[col for col in df.columns if col in selected_properties]]
                all_columns.update(df.columns)
                combined_data.append(df)

    # Ensure that there is data to combine
    if combined_data:
        # Ensure that all columns are included in the final dataframe
        final_columns = sorted(all_columns)
        final_df = pd.concat(combined_data, ignore_index=True, sort=False)
        final_df = final_df.reindex(columns=final_columns)
    else:
        print("No valid data found. Creating an empty validation file.")
        final_df = pd.DataFrame(columns=all_columns)

    # Define the validation Excel file path with timestamp
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    validation_file = os.path.join(output_dir, f"validation_output_{timestamp}.xlsx")
    print(f"Saving validation file to: {validation_file}")

    try:
        # Save the dataframe to Excel
        final_df.to_excel(validation_file, index=False)

        # Validate the Excel and apply any necessary formatting
        validate_excel(validation_file)
        
        print(f"Validation file created at: {validation_file}")

        # Show a clickable link to open the Excel file
        messagebox.showinfo(
            "Processing Complete",
            f"Processing complete!\n\nClick 'OK' to open the validation file.",
        )
        webbrowser.open(f"file://{validation_file}")
    except Exception as e:
        print(f"Error creating Excel: {e}")
        messagebox.showerror("Error", f"Error creating Excel: {e}")


# Validate Excel and apply error highlighting
def validate_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Get header row and map column names to their positions
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    try:
        idx_CCILevel1ParentLocationID = header_row.index("CCILevel1ParentLocationID")
        idx_CCILevel1ParentTypeID = header_row.index("CCILevel1ParentTypeID")
        idx_CCILevel2ParentLocationID = header_row.index("CCILevel2ParentLocationID")
        idx_CCILevel2ParentTypeID = header_row.index("CCILevel2ParentTypeID")
        idx_CCILocationID = header_row.index("CCILocationID")
        idx_CCIMultiLevelLocationID = header_row.index("CCIMultiLevelLocationID")
        idx_CCIMultiLevelTypeID = header_row.index("CCIMultiLevelTypeID")
    except ValueError as e:
        print(f"Required columns not found in the sheet: {e}")
        return

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        CCILevel2ParentTypeID = str(row[idx_CCILevel2ParentTypeID].value)
        CCILevel1ParentTypeID = str(row[idx_CCILevel1ParentTypeID].value)
        expected_type_id = f"§{CCILevel2ParentTypeID}.{CCILevel1ParentTypeID}"
        if str(row[idx_CCIMultiLevelTypeID].value) != expected_type_id:
            row[idx_CCIMultiLevelTypeID].fill = red_fill

        CCILevel2ParentLocationID = str(row[idx_CCILevel2ParentLocationID].value)
        CCILevel1ParentLocationID = str(row[idx_CCILevel1ParentLocationID].value)
        CCILocationID = str(row[idx_CCILocationID].value)
        expected_location_id = f"+{CCILevel2ParentLocationID}.{CCILevel1ParentLocationID}.{CCILocationID}"
        if str(row[idx_CCIMultiLevelLocationID].value) != expected_location_id:
            row[idx_CCIMultiLevelLocationID].fill = red_fill

    wb.save(excel_path)


# Select input directory
def select_input_directory():
    input_dir = filedialog.askdirectory(title="Select Input Directory")
    input_dir_entry.delete(0, tk.END)
    input_dir_entry.insert(0, input_dir)


# Select output directory
def select_output_directory():
    output_dir = filedialog.askdirectory(title="Select Output Directory")
    output_dir_entry.delete(0, tk.END)
    output_dir_entry.insert(0, output_dir)


# Select parameter list file
def select_property_list_file():
    property_file_path = filedialog.askopenfilename(title="Select Property List File", filetypes=[("Text Files", "*.txt")])
    if property_file_path:
        load_selected_properties(property_file_path)
        property_file_label.config(text=f"Loaded properties from: {property_file_path}")


# Start processing
def start_processing():
    global selected_properties
    selected_properties = [param.get() for param in parameter_checkboxes if param.get()]

    input_dir = input_dir_entry.get()
    output_dir = output_dir_entry.get()

    if not input_dir or not output_dir:
        messagebox.showerror("Error", "Please select both input and output directories.")
        return

    if not selected_properties:
        messagebox.showerror("Error", "Please select at least one parameter to export.")
        return

    try:
        print("Starting IFC to CSV conversion...")
        progress_var.set(0)
        progress_label.config(text="Initializing file processing...")
        process_ifc_directory(input_dir, output_dir)
        create_combined_excel(output_dir)
        messagebox.showinfo("Success", "Processing complete! CSV files and validation sheet saved.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


# Create Tkinter app
app = tk.Tk()
app.title("IFC to CSV Converter - (CC) balaji.work")
app.geometry("600x700")

tk.Label(app, text="Created for Eastern Ring Road Project by Balaji Balagurusami babs@cowi.com for COWI A/S.", fg="blue").pack(pady=5)
tk.Label(app, text="License: Creative Commons 0 1.0 Universal", fg="blue").pack(pady=5)

def open_github():
    webbrowser.open_new("https://github.com/balajibalagurusami/python/")

github_label = tk.Label(app, text="GitHub Repo", fg="blue", cursor="hand2")
github_label.pack(pady=5)
github_label.bind("<Button-1>", lambda e: open_github())

# Input directory
tk.Label(app, text="Select Input Directory:").pack()
input_dir_entry = tk.Entry(app, width=60)
input_dir_entry.pack(pady=5)
tk.Button(app, text="Browse", command=select_input_directory).pack()

# Output directory
tk.Label(app, text="Select Output Directory:").pack(pady=5)
output_dir_entry = tk.Entry(app, width=60)
output_dir_entry.pack(pady=5)
tk.Button(app, text="Browse", command=select_output_directory).pack()

# Property list selection
tk.Button(app, text="Load Property List", command=select_property_list_file).pack(pady=5)
property_file_label = tk.Label(app, text="No property list loaded.")
property_file_label.pack(pady=5)

# Scrollable frame for parameters
scroll_frame = tk.Frame(app)
scroll_frame.pack(pady=10)

# Create canvas and scrollbar for scrolling
canvas = tk.Canvas(scroll_frame)
scrollbar = tk.Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
canvas.config(yscrollcommand=scrollbar.set)

scrollable_frame = tk.Frame(canvas)

# Create the window inside the canvas
canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

scrollbar.pack(side="right", fill="y")
canvas.pack(side="left", fill="both", expand=True)

def load_checkboxes():
    global parameter_checkboxes
    # Clear previous checkboxes
    for widget in scrollable_frame.winfo_children():
        widget.destroy()

    parameter_checkboxes = []

    for param in available_properties:
        var = tk.BooleanVar()
        cb = tk.Checkbutton(scrollable_frame, text=param, variable=var)
        cb.pack(anchor="w")
        parameter_checkboxes.append(var)

    # Update scroll region
    scrollable_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

    # Enable the Start Processing button after loading checkboxes
    start_button.config(state="normal")

# "Get Available Parameters" Button
tk.Button(app, text="Get Available Parameters", command=lambda: process_ifc_directory(input_dir_entry.get(), output_dir_entry.get())).pack(pady=5)

# Start button and progress bar
start_button = tk.Button(app, text="Start Processing", command=start_processing, state="disabled")
start_button.pack(pady=10)
progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(app, variable=progress_var, length=400)
progress_bar.pack(pady=5)
progress_label = tk.Label(app, text="")
progress_label.pack()

app.mainloop()
